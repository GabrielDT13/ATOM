from flask import Flask, render_template, request, redirect, url_for, session, flash, jsonify, Response, stream_with_context, abort, send_file
from pathlib import Path
from auth import check_user
from auth import load_users
from auth import add_user
from auth import DATA_FILE  # Si quieres modificar users.json
from werkzeug.utils import secure_filename
import subprocess, json, os, re, shutil, time
from openpyxl import load_workbook
import pandas as pd


app = Flask(__name__)
app.secret_key = 'your_secret_key'

# Carpeta raiz
#BASE_DIR = os.path.dirname(os.path.abspath(__file__))
BASE_DIR = Path(__file__).resolve().parent  # ruta absoluta del proyecto


# Carpeta base para proyectos
PROJECTS_DIR = Path(BASE_DIR) / "projects"
ALLOWED_EXTENSIONS = {'xlsx', 'xls', 'csv', 'txt'}  # extensiones permitidas (puedes ampliar)

def load_json(filename):
    path = os.path.join('data', filename)  # Ajusta la carpeta donde tengas los JSON
    with open(path, 'r', encoding='utf-8') as f:  # <- clave: encoding UTF-8
        return json.load(f)


def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def all_designs_ready(project_dir):
    """
    Retorna True si todas las subcarpetas del proyecto tienen al menos un archivo .html.
    """
    for entry in project_dir.iterdir():
        if entry.is_dir():
            html_files = list(entry.glob('*.html'))
            if not html_files:
                return False
    return True


@app.route('/')
def index():
    if 'username' in session:
        return redirect(url_for('dashboard'))
    else:
        return redirect(url_for('login'))

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']

        if check_user(username, password):
            session['username'] = username
            session['role'] = 'admin' if username == 'admin' else 'user'
            return redirect(url_for('dashboard'))
        else:
            flash('Usuario o contraseña incorrectos')

    return render_template('pages/login.html')


@app.route('/dashboard')
def dashboard():
    if 'username' not in session:
        return redirect(url_for('login'))
    return render_template(
        'layouts/dashboard_shell.html',
        username=session['username'],
        content_template='pages/dashboard/home.html'
    )



@app.route('/logout')
def logout():
    session.pop('username', None)
    return redirect(url_for('login'))


# API endpoints dinámicos
@app.route('/api/messages')
def get_messages():
    if 'username' not in session:
        return jsonify({"error": "No autorizado"}), 401
    return jsonify(load_json('messages.json'))



@app.route('/api/sidebar_left')
def get_sidebar_left():
    if 'username' not in session:
        return jsonify({"error": "No autorizado"}), 401

    data = load_json('sidebar_left.json')

    # Filtrar items admin_only si el usuario no es admin
    if session.get('role') != 'admin':
        data['items'] = [item for item in data['items'] if not item.get('admin_only', False)]

    return jsonify(data)  # Flask jsonify ya enviará UTF-8 automáticamente



@app.route('/api/sidebar_right')
def get_sidebar_right():
    if 'username' not in session:
        return jsonify({"error": "No autorizado"}), 401

    username = session['username']
    user_dir = os.path.join('projects', username)




    def build_tree(path, current_project=None, parent_path=''):
        print("DEBUG: build_tree path =", path)  # <--- depuración
        tree = []
        html_exists_in_folder = False  # acumulador

        try:
            for entry in sorted(os.listdir(path)):
                full_path = os.path.join(path, entry)
                rel_path = os.path.join(parent_path, entry) if parent_path else entry  # <-- path relativo

                if os.path.isdir(full_path):
                    # proyecto (nivel raíz) o subcarpeta (ej. diseños)
                    project_name = entry if current_project is None else current_project

                    child_tree, child_html_exists = build_tree(full_path, project_name, rel_path)

                    tree.append({
                        "name": entry,
                        "type": "folder",
                        "html_exists": child_html_exists,
                        "children": child_tree,
                        "path": rel_path,      # <-- path relativo de la carpeta
                        "project_name": project_name
                    })

                    if child_html_exists:
                        html_exists_in_folder = True

                else:
                    is_html = entry.lower().endswith(".html")
                    if is_html:
                        html_exists_in_folder = True

                    tree.append({
                        "name": entry,
                        "type": "file",
                        "username": username,
                        "project_name": current_project,
                        "path": rel_path      # <-- path relativo del archivo
                    })
        except FileNotFoundError:
            pass

        return tree, html_exists_in_folder



    tree, _ = build_tree(user_dir)
    print(tree)  # 👈 debug
    return jsonify({"title": "Mis Proyectos", "items": tree})



@app.route('/register', methods=['GET', 'POST'])
def register():
    # Solo admin puede acceder
    if session.get('role') != 'admin':
        flash("No tienes permiso para acceder a esta página")
        return redirect(url_for('dashboard'))

    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        email = request.form['email']

        success, message = add_user(username, password, email)

        if success:
            # Crear carpeta de proyectos asociada al nuevo usuario
            projects_root = Path(BASE_DIR) / 'projects'
            user_dir = projects_root / username
            try:
                user_dir.mkdir(parents=True, exist_ok=True)
            except Exception as e:
                flash(f"Usuario creado, pero no se pudo crear su carpeta de proyectos: {e}", "error")

        return render_template(
            "layouts/dashboard_shell.html",
            username=session['username'],
            content_template="pages/dashboard/register.html",
            message=message,
            success=success
        )

    return render_template(
        "layouts/dashboard_shell.html",
        username=session['username'],
        content_template="pages/dashboard/register.html"
    )


@app.route('/dashboard/register')
def dashboard_register():
    if 'username' not in session:
        return redirect(url_for('login'))

    if session.get('role') != 'admin':
        flash("No autorizado")
        return redirect(url_for('dashboard'))

    return render_template(
        'layouts/dashboard_shell.html',
        username=session['username'],
        content_template='pages/dashboard/register.html'
    )




@app.route('/dashboard/edit_users')
def dashboard_edit_users():
    if 'username' not in session:
        return redirect(url_for('login'))

    # Admin ve todos los usuarios, usuario normal solo su propio usuario
    users = load_users()
    if session.get('role') != 'admin':
        users = {session['username']: users[session['username']]}

    return render_template(
        'layouts/dashboard_shell.html',
        username=session['username'],
        content_template='pages/dashboard/edit_users.html',
        users=users
    )


@app.route('/dashboard/edit_user/<username>', methods=['GET', 'POST'])
def edit_user_form(username):
    if 'username' not in session:
        return redirect(url_for('login'))

    users = load_users()

    # Solo admin o el propio usuario pueden editar
    if session.get('role') != 'admin' and session['username'] != username:
        flash("No autorizado")
        return redirect(url_for('dashboard'))

    user = users.get(username)
    if not user:
        flash("Usuario no encontrado")
        return redirect(url_for('dashboard_edit_users'))

    if request.method == 'POST':
        new_username = request.form['username']
        new_email = request.form['email']
        new_password = request.form['password']

        # Si cambia el nombre de usuario, renombrar carpeta
        if new_username != username:
            projects_root = Path(BASE_DIR) / 'projects'
            old_dir = projects_root / username
            new_dir = projects_root / new_username
            if old_dir.exists():
                old_dir.rename(new_dir)
            # Actualizar clave en JSON
            users[new_username] = users.pop(username)
            username = new_username  # actualizar variable para seguir usando

        # Actualizar datos
        user = users[username]
        user['email'] = new_email
        if new_password:
            user['password'] = new_password

        # Guardar cambios
        with open(DATA_FILE, 'w', encoding='utf-8') as f:
            json.dump(users, f, indent=4, ensure_ascii=False)

        flash(f"Usuario {username} actualizado correctamente", "success")
        return redirect(url_for('dashboard_edit_users'))

    # Renderizar dentro del dashboard, mismo estilo que registro
    return render_template(
        "layouts/dashboard_shell.html",
        username=session['username'],
        content_template="pages/dashboard/edit_user_form.html",
        user=user,
        username_form=username  # nombre actual del usuario para el form
    )




@app.route('/dashboard/delete_user/<username>', methods=['POST'])
def delete_user(username):
    # permisos: solo admin puede borrar usuarios
    if 'username' not in session or session.get('role') != 'admin':
        flash("No autorizado")
        return redirect(url_for('dashboard'))

    users = load_users()
    if username not in users:
        flash("El usuario no existe.", "error")
        return redirect(url_for('dashboard_edit_users'))

    # Eliminar usuario del JSON
    users.pop(username)
    with open(DATA_FILE, 'w', encoding='utf-8') as f:
        json.dump(users, f, indent=4, ensure_ascii=False)

    # Ruta absoluta segura a projects/<username>
    projects_root = Path(BASE_DIR) / 'projects'
    user_dir = projects_root / username

    try:
        if user_dir.exists() and user_dir.is_dir():
            # Seguridad: comprobar que user_dir está dentro de projects_root
            try:
                projects_root_resolved = projects_root.resolve()
                user_dir_resolved = user_dir.resolve()
            except Exception:
                # si resolve falla por permisos, evitar eliminar
                flash("No se pudo verificar la ruta de la carpeta del usuario. No se borró la carpeta.", "error")
                return redirect(url_for('dashboard_edit_users'))

            if str(user_dir_resolved).startswith(str(projects_root_resolved) + os.sep) or str(user_dir_resolved) == str(projects_root_resolved):
                shutil.rmtree(user_dir)  # elimina recursivamente
                flash(f"Usuario {username} y su carpeta de proyectos eliminados correctamente.", "success")
            else:
                flash("Ruta inválida: no se eliminó la carpeta del usuario por seguridad.", "error")
        else:
            flash(f"Usuario {username} eliminado. No existía carpeta de proyectos.", "success")
    except Exception as e:
        # Si hay error al borrar carpeta, lo informamos pero el usuario ya fue eliminado del JSON
        flash(f"Usuario eliminado, pero error al borrar carpeta: {e}", "error")

    return redirect(url_for('dashboard_edit_users'))



@app.route('/dashboard/create_project', methods=['GET', 'POST'])
def create_project():
    if 'username' not in session:
        return redirect(url_for('login'))

    username = session['username']
    user_dir = Path(BASE_DIR) / 'projects' / username

    # Asegurarse de que exista la carpeta del usuario
    user_dir.mkdir(parents=True, exist_ok=True)

    message = None
    success = False

    if request.method == 'POST':
        project_name = request.form['project_name']
        template_file = request.files.get('template_file')
        additional_files = request.files.getlist('additional_files')

        if not template_file:
            message = "Debes seleccionar un archivo Excel para el proyecto"
            success = False
        else:
            # Crear carpeta del proyecto dentro de la carpeta del usuario
            project_dir = user_dir / project_name
            if project_dir.exists():
                message = f"El proyecto '{project_name}' ya existe"
                success = False
            else:
                project_dir.mkdir()
                try:
                    # Guardar archivo Excel
                    excel_path = project_dir / template_file.filename
                    template_file.save(excel_path)

                    # Guardar archivos adicionales
                    for f in additional_files:
                        if f.filename:
                            f.save(project_dir / f.filename)

                    message = f"Proyecto '{project_name}' creado correctamente."
                    success = True
                except Exception as e:
                    message = f"Error al crear proyecto: {e}"
                    success = False

    return render_template(
        'layouts/dashboard_shell.html',
        username=username,
        content_template='pages/dashboard/create_project.html',
        message=message,
        success=success
    )


@app.route('/dashboard/edit_project/<username>/<project_name>', methods=['GET', 'POST'])
def edit_project_form(username, project_name):

    if 'username' not in session:
        return redirect(url_for('login'))

    # Permiso: solo admin o dueño del proyecto
    if session.get('role') != 'admin' and session['username'] != username:
        flash("No autorizado")
        return redirect(url_for('dashboard'))

    # Ruta del proyecto actual
    project_dir = Path(BASE_DIR) / 'projects' / username / project_name
    if not project_dir.exists() or not project_dir.is_dir():
        flash("Proyecto no encontrado", "error")
        return redirect(url_for('dashboard_edit_projects'))  # 👈 corregido

    if request.method == 'POST':
        new_name = request.form.get('project_name')
        excel_file = request.files.get('excel_file')
        additional_files = request.files.getlist('additional_files')

        # --- Renombrar proyecto si el nombre cambió ---
        if new_name and new_name != project_name:
            new_dir = Path(BASE_DIR) / 'projects' / username / new_name
            if new_dir.exists():
                flash("Ya existe un proyecto con ese nombre", "error")
                return redirect(url_for('edit_project_form', username=username, project_name=project_name))
            project_dir.rename(new_dir)
            project_dir = new_dir
            project_name = new_name

        # --- Reemplazar archivos si se suben nuevos ---
        if excel_file or (additional_files and any(f.filename for f in additional_files)):
            # borrar todos los archivos actuales dentro del proyecto
            for item in project_dir.iterdir():
                if item.is_file():
                    item.unlink()
                elif item.is_dir():
                    shutil.rmtree(item)

            # Guardar Excel (si hay)
            if excel_file and excel_file.filename:
                excel_path = project_dir / excel_file.filename
                excel_file.save(excel_path)

            # Guardar archivos adicionales
            for f in additional_files:
                if f and f.filename:
                    file_path = project_dir / f.filename
                    f.save(file_path)

        flash("Proyecto actualizado correctamente", "success")
        return redirect(url_for('dashboard_edit_projects'))  # 👈 corregido

    # Precargar datos: listar archivos del proyecto
    existing_files = []
    for item in project_dir.iterdir():
        if item.is_file():
            existing_files.append(item.name)

    return render_template(
        "layouts/dashboard_shell.html",
        content_template="pages/dashboard/edit_project_form.html",
        username=username,
        project_name=project_name,
        project_files=existing_files
    )


@app.route('/dashboard/edit_projects')
def dashboard_edit_projects():
    if 'username' not in session:
        return redirect(url_for('login'))

    username = session['username']
    projects_root = Path(BASE_DIR) / 'projects'

    # Admin ve todos los proyectos
    if session.get('role') == 'admin':
        users = [d.name for d in projects_root.iterdir() if d.is_dir()]
    else:
        users = [username]

    # Diccionario de proyectos: {usuario: [proyecto1, proyecto2,...]}
    projects = {}
    for user in users:
        user_dir = projects_root / user
        if user_dir.exists():
            projects[user] = [d.name for d in user_dir.iterdir() if d.is_dir()]
        else:
            projects[user] = []

    return render_template(
        'layouts/dashboard_shell.html',
        username=username,
        content_template='pages/dashboard/edit_projects.html',
        projects=projects
    )


@app.route('/dashboard/delete_project/<username>/<project_name>', methods=['POST'])
def delete_project(username, project_name):
    if 'username' not in session:
        return redirect(url_for('login'))

    # Solo admin o el propio usuario puede borrar su proyecto
    if session.get('role') != 'admin' and session['username'] != username:
        flash("No autorizado")
        return redirect(url_for('dashboard'))

    project_dir = Path(BASE_DIR) / 'projects' / username / project_name
    if project_dir.exists() and project_dir.is_dir():
        try:
            shutil.rmtree(project_dir)
            flash(f"Proyecto '{project_name}' eliminado correctamente.", "success")
        except Exception as e:
            flash(f"Error al eliminar el proyecto: {e}", "error")
    else:
        flash("Proyecto no encontrado.", "error")

    return redirect(url_for('dashboard_edit_projects'))




@app.route('/api/get_file_content_preview')
def get_file_content_preview():
    path = request.args.get('path')
    max_lines = int(request.args.get('max_lines', 30))

    if not path or not os.path.isfile(path):
        return "Archivo no encontrado", 404

    lines = []
    with open(path, 'r', encoding='utf-8') as f:
        for i, line in enumerate(f):
            if i >= max_lines:
                lines.append('...\n')  # indicamos que hay más contenido
                break
            lines.append(line)
    
    content = ''.join(lines)
    return content, 200, {'Content-Type': 'text/plain; charset=utf-8'}


@app.route('/view_file/<username>/<path:file_path>')
def view_file(username, file_path):
    base_path = Path('projects') / username
    full_path = base_path / file_path  # ahora es un objeto Path

    if not full_path.exists() or not full_path.is_file():
        return jsonify({"error": "Archivo no encontrado"}), 404

    # Leer contenido (por ejemplo)
    content = full_path.read_text(encoding='utf-8')
    return jsonify({"content": content})




@app.route('/run_rmd')
def run_rmd():

    #username = request.args.get("username")
    if 'username' not in session:
            return "No autorizado", 401
    username = session['username']            # usar siempre el usuario logueado

    project_name = request.args.get("project_name")
    if not project_name:
            return "Falta project_name", 400
 
    # Carpetas del proyecto
    project_dir = BASE_DIR / "projects" / username / project_name

    # Excel template.xlsx
    excel_file = project_dir / "template.xlsx"
    if not excel_file.exists():
        return "Archivo template.xlsx no encontrado", 404

    from openpyxl import load_workbook
    wb = load_workbook(excel_file)
    if "design" not in wb.sheetnames:
        return "Hoja 'design' no encontrada en template.xlsx", 400

    sheet = wb["design"]

    # Leer filas (asumiendo que la primera fila es header)
    headers = [cell for cell in next(sheet.iter_rows(values_only=True))]
    rows = list(sheet.iter_rows(min_row=2, values_only=True))  # desde la fila 2




    def generate():
        for row in rows:
            record = dict(zip(headers, row))
            designID = record.get("designID")
            analysis_type = record.get("analysis_type")

            if not designID or not analysis_type:
                yield f"data:Skipping row with missing designID or analysis_type\n\n"
                continue

            # Crear carpeta para designID
            design_dir = project_dir / designID
            design_dir.mkdir(parents=True, exist_ok=True)

            # Ruta del script Rmd
            rmd_file = BASE_DIR / "r_scripts" / f"{analysis_type}.Rmd"
            if not rmd_file.exists():
                yield f"data:Rmd file not found for analysis_type: {analysis_type}\n\n"
                continue

            # Comando Rscript
            output_file = design_dir / f"{designID}.html"
            cmd = [
                "Rscript",
                "-e",
                (
                    f'rmarkdown::render("{rmd_file}", '
                    f'output_file="{str(output_file)}", '
                    f'params=list(designID="{str(designID)}", base_dir="{str(BASE_DIR)}", project_dir="{str(project_dir)}", design_dir="{str(design_dir)}"))'
                )
            ]


            # 🔍 Mostrar comando antes de ejecutarlo
            print("=====================================")
            print("Ejecutando comando Rscript:")
            print(" ".join(cmd))
            print("=====================================")



            yield f"data:Running analysis for designID {designID} using {analysis_type}.Rmd\n\n"

            process = subprocess.Popen(
                cmd,
                stdout=subprocess.PIPE,
                stderr=subprocess.STDOUT,
                text=True,
                bufsize=1
            )

            for line in iter(process.stdout.readline, ''):
                if not line:
                    break
                yield f"data:{line.strip()}\n\n"

            process.wait()
            yield f"data:Finished analysis for designID {designID}\n\n"

            # --- limpieza de Resultados ---
            if os.path.exists(design_dir):
                try:
                    clean_resultados(design_dir, designID)
                    yield f"data:Cleaned {designID} folder, kept only HTML/ZIP/Excel\n\n"
                except Exception as e:
                    yield f"data:WARNING: Could not clean Resultados folder: {str(e)}\n\n"

        # 🔹 Espera hasta que todos los HTML estén generados antes de enviar ---FIN---
        while not all_designs_ready(project_dir):
            time.sleep(1)

        yield "data:---FIN---\n\n"        

    return Response(stream_with_context(generate()), mimetype='text/event-stream')


def clean_resultados(out_dir, design_id):
    """
    Borra todos los archivos/carpetas en out_dir excepto el HTML, ZIP y Excel.
    """
    keep_ext = {".html", ".zip", ".xlsx", ".docx", ".Rmd"}
    keep_files = {
        os.path.join(out_dir, f"{design_id}.html"),
        os.path.join(out_dir, f"{design_id}.zip"),
        os.path.join(out_dir, f"{design_id}.xlsx"),
        os.path.join(out_dir, f"{design_id}.docx"),
        os.path.join(out_dir, f"{design_id}.Rmd")
    }

    for item in os.listdir(out_dir):
        item_path = os.path.join(out_dir, item)
        _, ext = os.path.splitext(item)

        if item_path in keep_files or ext in keep_ext:
            continue

        if os.path.isdir(item_path):
            shutil.rmtree(item_path)
        else:
            os.remove(item_path)

@app.route('/projects/<username>/<path:filepath>')
def serve_project_file(username, filepath):
    user_dir = os.path.join('projects', username)
    abs_path = os.path.join(user_dir, filepath)
    if not os.path.exists(abs_path) or not os.path.isfile(abs_path):
        return "Not Found", 404
    return send_file(abs_path, as_attachment=True)


if __name__ == '__main__':
    host = os.getenv("FLASK_HOST", "0.0.0.0")
    port = int(os.getenv("FLASK_PORT", "5000"))
    debug = os.getenv("FLASK_ENV", "development") == "development"
    app.run(host=host, port=port, debug=debug)
